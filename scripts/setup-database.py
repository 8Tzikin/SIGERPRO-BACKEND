#!/usr/bin/env python3
"""
SIGER PRO Database Setup Script
Creates all tables and imports initial data
"""

import os
import sys
import mysql.connector
from mysql.connector import Error
from urllib.parse import urlparse
import openpyxl
from datetime import datetime

# Parse DATABASE_URL
db_url = os.getenv('DATABASE_URL')
if not db_url:
    print("❌ Error: DATABASE_URL not set")
    sys.exit(1)

# Parse connection string: mysql://user:password@host:port/database
parsed = urlparse(db_url)
config = {
    'host': parsed.hostname,
    'user': parsed.username,
    'password': parsed.password,
    'database': parsed.path.lstrip('/'),
    'port': parsed.port or 3306,
    'ssl_disabled': False,
}

# Add SSL configuration for TiDB
if 'tidb' in parsed.hostname.lower():
    config['ssl_verify_cert'] = False
    config['ssl_verify_identity'] = False

print("🚀 SIGER PRO Database Setup")
print("=" * 50)

# Connect to database
try:
    print("\n📊 Connecting to database...")
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()
    print("✓ Connected successfully")
except Error as err:
    print(f"❌ Connection error: {err}")
    sys.exit(1)

# Create tables
print("\n📝 Creating tables...")

tables = [
    # Roles table
    """
    CREATE TABLE IF NOT EXISTS roles (
        id INT PRIMARY KEY AUTO_INCREMENT,
        nombre VARCHAR(50) NOT NULL UNIQUE,
        descripcion TEXT,
        createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    
    # Users table
    """
    CREATE TABLE IF NOT EXISTS usuarios (
        id INT PRIMARY KEY AUTO_INCREMENT,
        openId VARCHAR(64) UNIQUE NOT NULL,
        nombre VARCHAR(255),
        email VARCHAR(320),
        loginMethod VARCHAR(64),
        roleId INT DEFAULT 1,
        intentosFallidos INT DEFAULT 0,
        bloqueado BOOLEAN DEFAULT FALSE,
        createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        updatedAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
        lastSignedIn TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (roleId) REFERENCES roles(id)
    )
    """,
    
    # Severity table
    """
    CREATE TABLE IF NOT EXISTS severidad (
        id INT PRIMARY KEY,
        letra VARCHAR(1),
        nombre VARCHAR(50),
        descripcion TEXT,
        valor INT
    )
    """,
    
    # Classifications table
    """
    CREATE TABLE IF NOT EXISTS clasificaciones (
        id INT PRIMARY KEY,
        nombre VARCHAR(50),
        descripcion TEXT,
        rango_min INT,
        rango_max INT,
        color VARCHAR(7)
    )
    """,
    
    # Normativa table
    """
    CREATE TABLE IF NOT EXISTS normativa (
        id INT PRIMARY KEY AUTO_INCREMENT,
        codigo VARCHAR(50) UNIQUE NOT NULL,
        nombre VARCHAR(255) NOT NULL,
        descripcion TEXT,
        tipo VARCHAR(50),
        createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    
    # Proveedores table
    """
    CREATE TABLE IF NOT EXISTS proveedores (
        id INT PRIMARY KEY AUTO_INCREMENT,
        nombre VARCHAR(255) NOT NULL,
        contacto VARCHAR(255),
        email VARCHAR(320),
        telefono VARCHAR(20),
        direccion TEXT,
        createdAt TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )
    """,
    
    # Evaluaciones table
    """
    CREATE TABLE IF NOT EXISTS evaluaciones (
        id INT PRIMARY KEY AUTO_INCREMENT,
        numeroReporte VARCHAR(100) UNIQUE NOT NULL,
        fecha TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        formaReporte VARCHAR(100),
        reportadoPor VARCHAR(255),
        entidadOrigen VARCHAR(255),
        descripcion TEXT NOT NULL,
        tipoOcurrencia VARCHAR(100),
        regulacion VARCHAR(100),
        probabilidad INT NOT NULL,
        severidadId INT NOT NULL,
        impacto INT,
        riesgoInherente INT,
        clasificacionId INT,
        estado VARCHAR(20) DEFAULT 'ABIERTO',
        creadoPor INT NOT NULL,
        fechaCreacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (creadoPor) REFERENCES usuarios(id),
        FOREIGN KEY (severidadId) REFERENCES severidad(id),
        FOREIGN KEY (clasificacionId) REFERENCES clasificaciones(id),
        INDEX idx_estado (estado),
        INDEX idx_fecha (fecha),
        INDEX idx_clasificacion (clasificacionId)
    )
    """,
    
    # Seguimiento table
    """
    CREATE TABLE IF NOT EXISTS seguimiento (
        id INT PRIMARY KEY AUTO_INCREMENT,
        evaluacionId INT NOT NULL,
        comentario TEXT,
        accion VARCHAR(255),
        usuarioId INT NOT NULL,
        fechaAccion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (evaluacionId) REFERENCES evaluaciones(id),
        FOREIGN KEY (usuarioId) REFERENCES usuarios(id)
    )
    """,
    
    # Bitacora table
    """
    CREATE TABLE IF NOT EXISTS bitacora (
        id INT PRIMARY KEY AUTO_INCREMENT,
        usuarioId INT NOT NULL,
        accion VARCHAR(50),
        tabla VARCHAR(50),
        registroId INT,
        detalles TEXT,
        fechaAccion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (usuarioId) REFERENCES usuarios(id),
        INDEX idx_fecha (fechaAccion)
    )
    """,
]

for table_sql in tables:
    try:
        cursor.execute(table_sql)
        print(f"✓ Table created")
    except Error as err:
        if "already exists" in str(err):
            print(f"✓ Table already exists")
        else:
            print(f"⚠️  Error: {err}")

conn.commit()

# Insert reference data
print("\n📥 Inserting reference data...")

# Insert roles
try:
    cursor.execute("DELETE FROM roles")
    cursor.execute("""
        INSERT INTO roles (id, nombre, descripcion) VALUES
        (1, 'SUPER_ADMIN', 'Super administrador con acceso total'),
        (2, 'ADMINISTRADOR', 'Administrador con acceso a gestión'),
        (3, 'AUDITOR', 'Auditor con acceso de solo lectura')
    """)
    conn.commit()
    print("✓ Roles inserted")
except Error as err:
    print(f"⚠️  Roles: {err}")

# Insert severity levels
try:
    cursor.execute("DELETE FROM severidad")
    cursor.execute("""
        INSERT INTO severidad (id, letra, nombre, descripcion, valor) VALUES
        (1, 'A', 'Muy Baja', 'Insignificante', 1),
        (2, 'B', 'Baja', 'Menor', 2),
        (3, 'C', 'Moderada', 'Moderada', 3),
        (4, 'D', 'Alta', 'Mayor', 4),
        (5, 'E', 'Crítica', 'Crítica', 5)
    """)
    conn.commit()
    print("✓ Severity levels inserted")
except Error as err:
    print(f"⚠️  Severity: {err}")

# Insert classifications
try:
    cursor.execute("DELETE FROM clasificaciones")
    cursor.execute("""
        INSERT INTO clasificaciones (id, nombre, descripcion, rango_min, rango_max, color) VALUES
        (1, 'BAJO', 'Riesgo bajo', 1, 5, '#70AD47'),
        (2, 'MEDIO', 'Riesgo medio', 6, 10, '#FFC000'),
        (3, 'ALTO', 'Riesgo alto', 11, 15, '#FF6B6B'),
        (4, 'CRÍTICO', 'Riesgo crítico', 16, 20, '#8B0000')
    """)
    conn.commit()
    print("✓ Classifications inserted")
except Error as err:
    print(f"⚠️  Classifications: {err}")

# Import Excel data
print("\n📄 Importing Excel data...")
excel_file = '/home/ubuntu/upload/EVALUACIONESRIESGOAGAREPORTES2025.xlsx'

if os.path.exists(excel_file):
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['MATRIZ']
        
        imported = 0
        skipped = 0
        errors = 0
        
        # Get headers
        headers = {}
        for col, cell in enumerate(ws[1], 1):
            if cell.value:
                headers[cell.value] = col
        
        print(f"  Found {len(headers)} columns")
        
        # Import rows
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), 2):
            try:
                # Extract values
                numero = None
                fecha = datetime.now()
                descripcion = ""
                probabilidad = 1
                severidad_id = 1
                
                for header, col in headers.items():
                    cell = row[col - 1]
                    value = cell.value
                    
                    if not value:
                        continue
                    
                    header_upper = str(header).upper()
                    
                    if 'NO.' in header_upper or 'NUMERO' in header_upper:
                        numero = str(value).strip()
                    elif 'FECHA' in header_upper:
                        if isinstance(value, datetime):
                            fecha = value
                    elif 'REPORTE' in header_upper and 'TEXTO' in header_upper:
                        descripcion = str(value).strip()[:500]
                    elif 'PROBABILIDAD' in header_upper:
                        try:
                            prob_str = str(value).upper().strip()
                            if prob_str in ['1', 'REMOTO']:
                                probabilidad = 1
                            elif prob_str in ['2', 'OCASIONAL']:
                                probabilidad = 2
                            elif prob_str in ['3', 'PROBABLE']:
                                probabilidad = 3
                            elif prob_str in ['4', 'FRECUENTE']:
                                probabilidad = 4
                        except:
                            pass
                    elif 'SEVERIDAD' in header_upper:
                        try:
                            sev_str = str(value).upper().strip()
                            if sev_str in ['A', '1', 'MUY BAJA']:
                                severidad_id = 1
                            elif sev_str in ['B', '2', 'BAJA']:
                                severidad_id = 2
                            elif sev_str in ['C', '3', 'MODERADA']:
                                severidad_id = 3
                            elif sev_str in ['D', '4', 'ALTA']:
                                severidad_id = 4
                            elif sev_str in ['E', '5', 'CRÍTICA']:
                                severidad_id = 5
                        except:
                            pass
                
                if not numero or not descripcion:
                    skipped += 1
                    continue
                
                # Calculate risk
                riesgo_inherente = probabilidad * severidad_id
                if riesgo_inherente <= 5:
                    clasificacion_id = 1
                elif riesgo_inherente <= 10:
                    clasificacion_id = 2
                elif riesgo_inherente <= 15:
                    clasificacion_id = 3
                else:
                    clasificacion_id = 4
                
                # Check if exists
                cursor.execute("SELECT id FROM evaluaciones WHERE numeroReporte = %s", (numero,))
                if cursor.fetchone():
                    skipped += 1
                    continue
                
                # Insert
                cursor.execute("""
                    INSERT INTO evaluaciones (
                        numeroReporte, fecha, descripcion, probabilidad, severidadId,
                        riesgoInherente, clasificacionId, estado, creadoPor
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (
                    numero, fecha, descripcion, probabilidad, severidad_id,
                    riesgo_inherente, clasificacion_id, 'ABIERTO', 1
                ))
                
                imported += 1
                
            except Exception as e:
                errors += 1
                print(f"  ⚠️  Row {row_idx}: {str(e)[:50]}")
        
        conn.commit()
        print(f"✓ Imported: {imported} records")
        print(f"⊘ Skipped: {skipped} records")
        if errors > 0:
            print(f"✗ Errors: {errors} records")
        
    except Exception as e:
        print(f"❌ Error importing Excel: {e}")
else:
    print(f"⚠️  Excel file not found: {excel_file}")

# Verify
print("\n✅ Verification")
print("=" * 50)

try:
    cursor.execute("SELECT COUNT(*) FROM usuarios")
    user_count = cursor.fetchone()[0]
    print(f"✓ Users: {user_count}")
    
    cursor.execute("SELECT COUNT(*) FROM evaluaciones")
    eval_count = cursor.fetchone()[0]
    print(f"✓ Evaluations: {eval_count}")
    
    cursor.execute("SELECT COUNT(*) FROM severidad")
    sev_count = cursor.fetchone()[0]
    print(f"✓ Severity levels: {sev_count}")
    
    cursor.execute("SELECT COUNT(*) FROM clasificaciones")
    class_count = cursor.fetchone()[0]
    print(f"✓ Classifications: {class_count}")
    
except Error as err:
    print(f"❌ Verification error: {err}")

cursor.close()
conn.close()

print("\n✅ Database setup completed successfully!\n")
