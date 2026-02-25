#!/usr/bin/env python3
"""
SIGER PRO Excel Importer - Corrected
Imports evaluation data from Excel MATRIZ sheet (row 3 has headers)
"""

import os
import sys
import mysql.connector
from mysql.connector import Error
from urllib.parse import urlparse
import openpyxl
from datetime import datetime

# Parse DATABASE_URL
db_url = os.getenv('DATABASE_URL')
if not db_url:
    print("❌ Error: DATABASE_URL not set")
    sys.exit(1)

parsed = urlparse(db_url)
config = {
    'host': parsed.hostname,
    'user': parsed.username,
    'password': parsed.password,
    'database': parsed.path.lstrip('/'),
    'port': parsed.port or 3306,
    'ssl_disabled': False,
}

print("🚀 SIGER PRO Excel Importer")
print("=" * 50)

# Connect
try:
    print("\n📊 Connecting to database...")
    conn = mysql.connector.connect(**config)
    cursor = conn.cursor()
    print("✓ Connected")
except Error as err:
    print(f"❌ Connection error: {err}")
    sys.exit(1)

# Load Excel
excel_file = '/home/ubuntu/upload/EVALUACIONESRIESGOAGAREPORTES2025.xlsx'

if not os.path.exists(excel_file):
    print(f"❌ File not found: {excel_file}")
    sys.exit(1)

print(f"\n📄 Loading Excel: {excel_file}")
wb = openpyxl.load_workbook(excel_file)
ws = wb['MATRIZ']

print(f"✓ Sheet 'MATRIZ' loaded")
print(f"✓ Max row: {ws.max_row}, Max col: {ws.max_column}")

# Headers are in row 3
header_row = 3
headers = {}

print(f"\n📋 Reading headers from row {header_row}...")
for col_idx, cell in enumerate(ws[header_row], 1):
    if cell.value:
        header_str = str(cell.value).strip().upper()
        headers[header_str] = col_idx
        if len(headers) <= 15:  # Print first 15
            print(f"  Col {col_idx}: {header_str[:50]}")

print(f"✓ Found {len(headers)} columns")

# Import data starting from row 4
print("\n📥 Importing data from row 4...")
imported = 0
skipped = 0
errors = 0

for row_idx in range(4, ws.max_row + 1):
    try:
        row = ws[row_idx]
        
        # Extract values
        numero = None
        fecha = datetime.now()
        forma_reporte = None
        reportado_por = None
        entidad_origen = None
        descripcion = ""
        tipo_ocurrencia = None
        regulacion = None
        probabilidad = 1
        severidad_id = 1
        
        # Map columns
        for header_key, col_idx in headers.items():
            cell = row[col_idx - 1]  # Adjust for 0-based indexing
            value = cell.value
            
            if not value:
                continue
            
            # Skip formulas
            if isinstance(value, str) and value.startswith('='):
                continue
            
            value_str = str(value).strip()
            
            if 'NO.' in header_key:
                numero = value_str
            elif 'FECHA' in header_key:
                if isinstance(value, datetime):
                    fecha = value
            elif 'FORMA' in header_key and 'REPORTE' in header_key:
                forma_reporte = value_str[:100]
            elif 'PRESENTADO' in header_key and 'POR' in header_key:
                reportado_por = value_str[:255]
            elif 'ENTIDAD' in header_key and 'ORIGEN' in header_key:
                entidad_origen = value_str[:255]
            elif 'REPORTE' in header_key and 'TEXTO' in header_key:
                descripcion = value_str[:1000]
            elif 'TIPO' in header_key and 'OCURRENCIA' in header_key:
                tipo_ocurrencia = value_str[:100]
            elif 'REGULACIÓN' in header_key or 'REGULACION' in header_key:
                regulacion = value_str[:100]
            elif 'PROBABILIDAD' in header_key:
                try:
                    prob_str = value_str.upper()
                    if prob_str in ['1', 'REMOTO']:
                        probabilidad = 1
                    elif prob_str in ['2', 'OCASIONAL']:
                        probabilidad = 2
                    elif prob_str in ['3', 'PROBABLE']:
                        probabilidad = 3
                    elif prob_str in ['4', 'FRECUENTE']:
                        probabilidad = 4
                    else:
                        probabilidad = max(1, min(4, int(float(prob_str))))
                except:
                    probabilidad = 1
            elif 'SEVERIDAD' in header_key:
                try:
                    sev_str = value_str.upper()
                    if sev_str in ['A', '1']:
                        severidad_id = 1
                    elif sev_str in ['B', '2']:
                        severidad_id = 2
                    elif sev_str in ['C', '3']:
                        severidad_id = 3
                    elif sev_str in ['D', '4']:
                        severidad_id = 4
                    elif sev_str in ['E', '5']:
                        severidad_id = 5
                    else:
                        severidad_id = max(1, min(5, int(float(sev_str))))
                except:
                    severidad_id = 1
        
        # Validate
        if not numero:
            skipped += 1
            continue
        
        if not descripcion or len(descripcion) < 5:
            descripcion = f"Reporte {numero}"
        
        # Calculate risk
        riesgo_inherente = probabilidad * severidad_id
        if riesgo_inherente <= 5:
            clasificacion_id = 1
        elif riesgo_inherente <= 10:
            clasificacion_id = 2
        elif riesgo_inherente <= 15:
            clasificacion_id = 3
        else:
            clasificacion_id = 4
        
        # Check if exists
        cursor.execute("SELECT id FROM evaluaciones WHERE numeroReporte = %s", (str(numero),))
        if cursor.fetchone():
            skipped += 1
            continue
        
        # Insert
        cursor.execute("""
            INSERT INTO evaluaciones (
                numeroReporte, fecha, formaReporte, reportadoPor, entidadOrigen,
                descripcion, tipoOcurrencia, regulacion, probabilidad, severidadId,
                riesgoInherente, clasificacionId, estado, creadoPor
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            str(numero), fecha, forma_reporte, reportado_por, entidad_origen,
            descripcion, tipo_ocurrencia, regulacion, probabilidad, severidad_id,
            riesgo_inherente, clasificacion_id, 'ABIERTO', 1
        ))
        
        imported += 1
        if imported % 50 == 0:
            print(f"  {imported} records imported...")
        
    except Exception as e:
        errors += 1
        if errors <= 5:
            print(f"  ⚠️  Row {row_idx}: {str(e)[:80]}")

conn.commit()

print(f"\n✅ Import Summary")
print("=" * 50)
print(f"✓ Imported: {imported} records")
print(f"⊘ Skipped: {skipped} records")
if errors > 0:
    print(f"✗ Errors: {errors} records")

# Verify
try:
    cursor.execute("SELECT COUNT(*) FROM evaluaciones")
    total = cursor.fetchone()[0]
    print(f"\n✓ Total evaluations in database: {total}")
    
    # Show distribution
    cursor.execute("""
        SELECT clasificacionId, COUNT(*) as count FROM evaluaciones 
        GROUP BY clasificacionId ORDER BY clasificacionId
    """)
    print(f"\n✓ Distribution by classification:")
    for class_id, count in cursor.fetchall():
        names = {1: 'BAJO', 2: 'MEDIO', 3: 'ALTO', 4: 'CRÍTICO'}
        print(f"  {names.get(class_id, 'Unknown')}: {count}")
except Exception as e:
    print(f"⚠️  Verification error: {e}")

cursor.close()
conn.close()

print("\n✅ Import completed!\n")
